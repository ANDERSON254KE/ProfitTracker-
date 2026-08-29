import pandas as pd
import openpyxl
from pathlib import Path

BASE = Path(r"C:\Users\Lenovo\OneDrive\Desktop\Projects\Profit Tracker")

price = pd.read_excel(BASE / "price list.xlsx", sheet_name="all")
price.columns = [str(c).strip() for c in price.columns]

wb = openpyxl.load_workbook(BASE / "stck sheet.xlsx")
ws = wb.active

# Locate header row (cell containing 'ITEM') and the first summary row.
header_row = None
summary_row = None
for r in range(1, ws.max_row + 1):
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=r, column=c).value
        if v is not None and str(v).strip().upper() == "ITEM":
            header_row = r
        if v is not None and "BALANCE CARRIED FORWARD" in str(v).upper():
            summary_row = r
            break
    if summary_row is not None:
        break

assert header_row and summary_row, f"header={header_row} summary={summary_row}"

# Remove existing item rows (between header and the summary block).
ws.delete_rows(header_row + 1, summary_row - header_row - 1)

# Build new item rows from the price list, grouped by category.
new_rows = []
last_cat = object()  # sentinel so first real category always emits a header
no = 1
for _, row in price.iterrows():
    name = row.get("Product")
    if name is None or (isinstance(name, float) and pd.isna(name)):
        continue
    name = str(name).strip()
    if name == "":
        continue
    cat = row.get("Category")
    cat = str(cat).strip() if not (cat is None or (isinstance(cat, float) and pd.isna(cat))) else ""
    price_val = row.get("Selling Price")
    selling = None
    if not (price_val is None or (isinstance(price_val, float) and pd.isna(price_val))):
        try:
            selling = float(price_val)
        except (TypeError, ValueError):
            selling = None

    if cat and cat != last_cat:
        new_rows.append([None, cat.upper(), None, None, None, None, None, None, None])
        last_cat = cat
    elif not cat:
        last_cat = object()  # force a header next time a real category appears

    new_rows.append([no, name, None, None, None, selling, None, None, None])
    no += 1

# Insert the new rows starting right after the header.
insert_at = header_row + 1
ws.insert_rows(insert_at, len(new_rows))
for i, vals in enumerate(new_rows):
    r = insert_at + i
    for c, v in enumerate(vals, start=1):
        ws.cell(row=r, column=c).value = v

wb.save(BASE / "stck sheet.xlsx")
print(f"Rebuilt {len(new_rows)} item rows. Summary now at row {insert_at + len(new_rows)}")
